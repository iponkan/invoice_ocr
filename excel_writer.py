from __future__ import annotations

from datetime import datetime
from pathlib import Path
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from config import EXCEL_COLUMNS


DEFAULT_COLUMN_WIDTH = (10, 45)
COLUMN_WIDTH_LIMITS = {
    "名称": (12, 32),
    "交款人": (18, 45),
    "项目名称": (18, 60),
    "金额": (10, 14),
    "申请号": (16, 24),
    "缴费日期": (14, 18),
    "处理状态": (10, 14),
    "错误信息": (18, 80),
}


def write_excel(records: list[dict[str, str]], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    last_error: PermissionError | None = None
    for candidate in _output_candidates(path):
        try:
            _write_workbook(records, candidate)
            _adjust_column_width(candidate)
            return candidate
        except PermissionError as exc:
            last_error = exc

    raise PermissionError(
        f"无法写入Excel文件，可能文件正在被打开: {path}"
    ) from last_error


def _write_workbook(records: list[dict[str, str]], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(EXCEL_COLUMNS)

    for record in records:
        sheet.append([record.get(column, "") for column in EXCEL_COLUMNS])

    workbook.save(path)


def _output_candidates(path: Path) -> list[Path]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return [
        path,
        path.with_name(f"{path.stem}_{timestamp}{path.suffix}"),
        path.with_name(f"{path.stem}_{timestamp}_2{path.suffix}"),
    ]


def _adjust_column_width(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        header = "" if column_cells[0].value is None else str(column_cells[0].value)
        min_width, max_width = COLUMN_WIDTH_LIMITS.get(header, DEFAULT_COLUMN_WIDTH)
        max_width_seen = _display_width(header)
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_width_seen = max(max_width_seen, _display_width(value))
            cell.alignment = Alignment(vertical="center", wrap_text=header == "错误信息")

        sheet.column_dimensions[column_letter].width = min(
            max(max_width_seen + 2, min_width),
            max_width,
        )

    workbook.save(path)


def _display_width(value: str) -> int:
    width = 0
    for char in value:
        width += 2 if unicodedata.east_asian_width(char) in {"F", "W"} else 1
    return width
